# export_fee_gate_research_v2.py
# Exports:
#   1) coin_level -> one row per token
#   2) ping_level -> one row per ping
#
# Pulls data from:
#   - alerts
#   - tokens
#   - fee_gate_log
#
# No pandas required.

import sqlite3
from pathlib import Path
from datetime import datetime
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

DB_PATH = Path("data/bot.db")
OUTPUT_XLSX = Path("fee_gate_research_export.xlsx")

# Set to the first time fee gate started being logged
START_DATETIME = "2026-04-09 18:43:00"

def to_unix(dt_str: str) -> float:
    return datetime.strptime(dt_str, "%Y-%m-%d %H:%M:%S").timestamp()

def normalize_time(ts):
    if ts is None:
        return None
    try:
        return datetime.fromtimestamp(float(ts)).strftime("%Y-%m-%d %H:%M:%S")
    except:
        return str(ts)

def autofit_worksheet(ws):
    for col_cells in ws.columns:
        max_length = 0
        col_letter = get_column_letter(col_cells[0].column)
        for cell in col_cells:
            try:
                value = "" if cell.value is None else str(cell.value)
                if len(value) > max_length:
                    max_length = len(value)
            except:
                pass
        ws.column_dimensions[col_letter].width = min(max_length + 2, 42)

def write_sheet(ws, rows, headers):
    ws.append(headers)
    for row in rows:
        ws.append([row.get(h) for h in headers])
    autofit_worksheet(ws)

def main():
    if not DB_PATH.exists():
        raise FileNotFoundError(f"Database not found: {DB_PATH}")

    start_ts = to_unix(START_DATETIME)

    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    cur = conn.cursor()

    # Pull fee-gated alert rows and join whatever token state exists
    rows = cur.execute("""
        SELECT
            f.token_address,
            f.symbol,
            f.alert_tier,
            f.tier_name,
            f.alert_time,

            f.total_fee,
            f.lp_fee,
            f.proto_fee,
            f.creator_fee,
            f.rate,
            f.events,
            f.creator_share,
            f.proto_share,
            f.fee_per_event,
            f.proto_to_lp,
            f.score,
            f.flags,
            f.label,
            f.manual_verdict,
            f.reviewed_at,

            a.alert_price,
            a.alert_mcap,
            a.ath_price AS alert_ath_price,
            a.ath_mcap AS alert_ath_mcap,
            a.peak_price_after,
            a.peak_mcap_after,

            t.pool_address,
            t.status,
            t.migration_price,
            t.migration_mcap,
            t.current_price,
            t.current_mcap,
            t.liquidity_usd,
            t.ath_price AS token_ath_price,
            t.ath_mcap AS token_ath_mcap,
            t.ath_time,
            t.volume_1h,
            t.volume_6h,
            t.volume_24h,
            t.migration_time,
            t.last_price_update,
            t.last_alerted_tier

        FROM fee_gate_log f
        LEFT JOIN alerts a
            ON a.address = f.token_address
           AND a.alert_time = f.alert_time
           AND a.tier_index = f.alert_tier
        LEFT JOIN tokens t
            ON t.address = f.token_address
        WHERE f.alert_time >= ?
        ORDER BY f.alert_time ASC
    """, (start_ts,)).fetchall()

    if not rows:
        print("No fee_gate_log rows found after cutoff.")
        print("Check START_DATETIME.")
        return

    ping_rows = []
    grouped = {}

    for r in rows:
        tier_num = None
        if r["alert_tier"] is not None:
            tier_num = int(r["alert_tier"]) + 1

        record = {
            "alert_time": normalize_time(r["alert_time"]),
            "token_name": r["symbol"],
            "token_address": r["token_address"],

            "tier_num": tier_num,
            "tier_label": r["tier_name"],

            # alert snapshot
            "alert_price": r["alert_price"],
            "alert_mcap": r["alert_mcap"],
            "alert_ath_price": r["alert_ath_price"],
            "alert_ath_mcap": r["alert_ath_mcap"],
            "peak_price_after": r["peak_price_after"],
            "peak_mcap_after": r["peak_mcap_after"],

            # token state
            "pool_address": r["pool_address"],
            "status": r["status"],
            "migration_price": r["migration_price"],
            "migration_mcap": r["migration_mcap"],
            "current_price": r["current_price"],
            "current_mcap": r["current_mcap"],
            "liquidity_usd": r["liquidity_usd"],
            "token_ath_price": r["token_ath_price"],
            "token_ath_mcap": r["token_ath_mcap"],
            "ath_time": normalize_time(r["ath_time"]),
            "volume_1h": r["volume_1h"],
            "volume_6h": r["volume_6h"],
            "volume_24h": r["volume_24h"],
            "migration_time": normalize_time(r["migration_time"]),
            "last_price_update": normalize_time(r["last_price_update"]),
            "last_alerted_tier": r["last_alerted_tier"],

            # fee gate
            "total_fee": r["total_fee"],
            "lp_fee": r["lp_fee"],
            "proto_fee": r["proto_fee"],
            "creator_fee": r["creator_fee"],
            "rate": r["rate"],
            "events": r["events"],
            "creator_share": r["creator_share"],
            "proto_share": r["proto_share"],
            "fee_per_event": r["fee_per_event"],
            "proto_to_lp": r["proto_to_lp"],
            "score": r["score"],
            "flags": r["flags"],
            "fee_gate_label": r["label"],
            "manual_verdict": r["manual_verdict"],
            "reviewed_at": normalize_time(r["reviewed_at"]),

            # manual research columns
            "manual_label": None,          # organic / scam / suspicious / stillborn
            "manual_result": None,         # 2x / 4x / dead / no bounce
            "manual_notes": None,
        }

        ping_rows.append(record)
        grouped.setdefault(r["token_address"], []).append(record)

    # sort ping rows
    ping_rows.sort(key=lambda x: ((x["alert_time"] or ""), (x["token_name"] or ""), (x["tier_num"] or 0)))

    coin_rows = []

    for token_address, group in grouped.items():
        group.sort(key=lambda x: x["alert_time"] or "")
        first = group[0]

        tiers_hit = sorted({g["tier_num"] for g in group if g["tier_num"] is not None})
        max_peak = max([g["peak_mcap_after"] or 0 for g in group], default=0)

        row = {
            "token_name": first["token_name"],
            "token_address": token_address,
            "pool_address": first["pool_address"],

            "first_alert_time": first["alert_time"],
            "migration_time": first["migration_time"],
            "ath_time": first["ath_time"],

            "status": first["status"],
            "migration_mcap": first["migration_mcap"],
            "first_alert_mcap": first["alert_mcap"],
            "token_ath_mcap": first["token_ath_mcap"],
            "liquidity_usd": first["liquidity_usd"],
            "volume_1h": first["volume_1h"],
            "volume_6h": first["volume_6h"],
            "volume_24h": first["volume_24h"],

            "max_tier_hit": max(tiers_hit) if tiers_hit else None,
            "tier_count": len(tiers_hit),
            "all_tiers_called": ", ".join([f"T{t}" for t in tiers_hit]) if tiers_hit else None,
            "peak_mcap_after_max": max_peak,

            # final/manual columns
            "manual_label": None,
            "manual_result": None,
            "manual_notes": None,
        }

        # first seen gate snapshot
        row["first_fee_gate_label"] = first["fee_gate_label"]
        row["first_score"] = first["score"]
        row["first_flags"] = first["flags"]
        row["first_total_fee"] = first["total_fee"]
        row["first_lp_fee"] = first["lp_fee"]
        row["first_proto_fee"] = first["proto_fee"]
        row["first_creator_fee"] = first["creator_fee"]
        row["first_rate"] = first["rate"]
        row["first_events"] = first["events"]
        row["first_creator_share"] = first["creator_share"]
        row["first_proto_share"] = first["proto_share"]
        row["first_fee_per_event"] = first["fee_per_event"]
        row["first_proto_to_lp"] = first["proto_to_lp"]

        for tier in [1, 2, 3]:
            tg = [g for g in group if g["tier_num"] == tier]
            if tg:
                t = tg[0]
                row[f"T{tier}_alert_time"] = t["alert_time"]
                row[f"T{tier}_alert_mcap"] = t["alert_mcap"]
                row[f"T{tier}_alert_ath_mcap"] = t["alert_ath_mcap"]
                row[f"T{tier}_peak_mcap_after"] = t["peak_mcap_after"]

                row[f"T{tier}_total_fee"] = t["total_fee"]
                row[f"T{tier}_lp_fee"] = t["lp_fee"]
                row[f"T{tier}_proto_fee"] = t["proto_fee"]
                row[f"T{tier}_creator_fee"] = t["creator_fee"]
                row[f"T{tier}_rate"] = t["rate"]
                row[f"T{tier}_events"] = t["events"]
                row[f"T{tier}_creator_share"] = t["creator_share"]
                row[f"T{tier}_proto_share"] = t["proto_share"]
                row[f"T{tier}_fee_per_event"] = t["fee_per_event"]
                row[f"T{tier}_proto_to_lp"] = t["proto_to_lp"]
                row[f"T{tier}_score"] = t["score"]
                row[f"T{tier}_flags"] = t["flags"]
                row[f"T{tier}_fee_gate_label"] = t["fee_gate_label"]
                row[f"T{tier}_manual_verdict"] = t["manual_verdict"]
            else:
                row[f"T{tier}_alert_time"] = None
                row[f"T{tier}_alert_mcap"] = None
                row[f"T{tier}_alert_ath_mcap"] = None
                row[f"T{tier}_peak_mcap_after"] = None

                row[f"T{tier}_total_fee"] = None
                row[f"T{tier}_lp_fee"] = None
                row[f"T{tier}_proto_fee"] = None
                row[f"T{tier}_creator_fee"] = None
                row[f"T{tier}_rate"] = None
                row[f"T{tier}_events"] = None
                row[f"T{tier}_creator_share"] = None
                row[f"T{tier}_proto_share"] = None
                row[f"T{tier}_fee_per_event"] = None
                row[f"T{tier}_proto_to_lp"] = None
                row[f"T{tier}_score"] = None
                row[f"T{tier}_flags"] = None
                row[f"T{tier}_fee_gate_label"] = None
                row[f"T{tier}_manual_verdict"] = None

        coin_rows.append(row)

    coin_rows.sort(key=lambda x: x["first_alert_time"] or "")

    ping_headers = [
        "alert_time", "token_name", "token_address", "tier_num", "tier_label",
        "alert_price", "alert_mcap", "alert_ath_price", "alert_ath_mcap",
        "peak_price_after", "peak_mcap_after",
        "pool_address", "status", "migration_price", "migration_mcap",
        "current_price", "current_mcap", "liquidity_usd",
        "token_ath_price", "token_ath_mcap", "ath_time",
        "volume_1h", "volume_6h", "volume_24h",
        "migration_time", "last_price_update", "last_alerted_tier",
        "total_fee", "lp_fee", "proto_fee", "creator_fee",
        "rate", "events", "creator_share", "proto_share",
        "fee_per_event", "proto_to_lp", "score", "flags",
        "fee_gate_label", "manual_verdict", "reviewed_at",
        "manual_label", "manual_result", "manual_notes"
    ]

    coin_headers = [
        "token_name", "token_address", "pool_address",
        "first_alert_time", "migration_time", "ath_time",
        "status", "migration_mcap", "first_alert_mcap", "token_ath_mcap",
        "liquidity_usd", "volume_1h", "volume_6h", "volume_24h",
        "max_tier_hit", "tier_count", "all_tiers_called", "peak_mcap_after_max",

        "first_fee_gate_label", "first_score", "first_flags",
        "first_total_fee", "first_lp_fee", "first_proto_fee", "first_creator_fee",
        "first_rate", "first_events", "first_creator_share", "first_proto_share",
        "first_fee_per_event", "first_proto_to_lp",

        "manual_label", "manual_result", "manual_notes",

        "T1_alert_time", "T1_alert_mcap", "T1_alert_ath_mcap", "T1_peak_mcap_after",
        "T1_total_fee", "T1_lp_fee", "T1_proto_fee", "T1_creator_fee",
        "T1_rate", "T1_events", "T1_creator_share", "T1_proto_share",
        "T1_fee_per_event", "T1_proto_to_lp", "T1_score", "T1_flags",
        "T1_fee_gate_label", "T1_manual_verdict",

        "T2_alert_time", "T2_alert_mcap", "T2_alert_ath_mcap", "T2_peak_mcap_after",
        "T2_total_fee", "T2_lp_fee", "T2_proto_fee", "T2_creator_fee",
        "T2_rate", "T2_events", "T2_creator_share", "T2_proto_share",
        "T2_fee_per_event", "T2_proto_to_lp", "T2_score", "T2_flags",
        "T2_fee_gate_label", "T2_manual_verdict",

        "T3_alert_time", "T3_alert_mcap", "T3_alert_ath_mcap", "T3_peak_mcap_after",
        "T3_total_fee", "T3_lp_fee", "T3_proto_fee", "T3_creator_fee",
        "T3_rate", "T3_events", "T3_creator_share", "T3_proto_share",
        "T3_fee_per_event", "T3_proto_to_lp", "T3_score", "T3_flags",
        "T3_fee_gate_label", "T3_manual_verdict",
    ]

    wb = Workbook()

    ws1 = wb.active
    ws1.title = "coin_level"
    write_sheet(ws1, coin_rows, coin_headers)

    ws2 = wb.create_sheet("ping_level")
    write_sheet(ws2, ping_rows, ping_headers)

    wb.save(OUTPUT_XLSX)
    print(f"Done. Exported {len(coin_rows)} coins and {len(ping_rows)} pings to {OUTPUT_XLSX}")

if __name__ == "__main__":
    main()